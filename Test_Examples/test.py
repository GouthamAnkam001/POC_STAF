import configparser
from openpyxl import load_workbook
from Common_functions.Appium.Android.android_keyword_mapping import KeywordMapping

key_word = KeywordMapping()

config = configparser.ConfigParser()
choice = int(input("enter input format \n 1 for ini \n 2 for Excel :"))
if choice == 1:
    STATUS = "ini"
    INI_PATH = "android_test_case/features/configuration/android/contacts.ini"


    def read_ini_file(ini_file_path, status):
        """

        :param ini_file_path: .ini file path to get locators
        :param status: whether it isa ini or excel
        :return:
        """
        config.read(ini_file_path)
        for section in config.sections():
            section_contents = dict(config[section])
            key_word.run_keywords(section_contents)

    read_ini_file(INI_PATH, STATUS)

if choice == 2:
    STATUS = "excel"
    PATH = r"/features/steps/Test_Examples/properties.ini"
    config.read(PATH)
    EXCEL = config['configuration']['excel_path']
    workbook = load_workbook(EXCEL)
    active_sheet = workbook.active

    for i in range(2, active_sheet.max_row + 1):
        section = []
        for j in range(1, active_sheet.max_column + 1):
            cell_obj = active_sheet.cell(row=i, column=j)
            section.append(cell_obj.value)
        key_word.run_keywords(section, STATUS)




#  "appium:platformName": "Android",
#         "appium:platformVersion": "10",
#         "appium:deviceName": "samsung",
#         "udid": "1bd50b18b10c7ece",
#         "appium:appPackage": "com.samsung.android.dialer",
#         "appium:appActivity": "com.samsung.android.dialer.DialtactsActivity"
#
#
# platform_UDID:1bd50b18b10c7ece
# app_package:com.samsung.android.dialer
# app_activity:com.samsung.android.dialer.DialtactsActivity